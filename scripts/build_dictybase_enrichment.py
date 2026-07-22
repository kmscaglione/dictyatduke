#!/usr/bin/env python3
"""Turn the mirrored dictyBase Downloads files into per-gene enrichment assets.

Reads the hardcopies under assets/dictybase-downloads/ (fetched by
fetch_dictybase_downloads.py) and writes:

  gene_extras.json        {DDB_G: {pmids, curation, transcripts, orthologs,
                           myristoylation, phospho, mw}}  — small per-gene facts
  dictybase_domains.json  {DDB_G: [InterPro/Pfam/etc. domains]}
  phenotype_ontology.json {term_lower: {id, definition, synonyms, parents}}
  codon_usage.json        {aa: {codon: fraction}}  — Dictyostelium codon bias

Many source files are keyed by the old DDB feature id, so we first build a
DDB -> DDB_G map from DDB-GeneID-UniProt.txt. Standard library + openpyxl (for
the one .xlsx). Run after fetch_dictybase_downloads.py:

  python3 scripts/build_dictybase_enrichment.py
"""
import csv
import json
import pathlib
import re
import ssl
import urllib.request

_SSL = ssl.create_default_context()
_SSL.check_hostname = False
_SSL.verify_mode = ssl.CERT_NONE
_UA = "dictyBase-data-sync/1.0 (+https://dicty.labs.duke.edu)"

ROOT = pathlib.Path(__file__).resolve().parents[1]
DL = ROOT / "assets" / "dictybase-downloads"
ASSETS = ROOT / "assets"
G = DL / "general"

DDBG = re.compile(r"DDB_G\d+")


def rows(path, skip_header=True):
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    for r in csv.reader(lines[1:] if skip_header else lines, delimiter="\t"):
        if r:
            yield r


def ddb_to_ddbg():
    """Old DDB feature id -> DDB_G gene id (from DDB-GeneID-UniProt.txt)."""
    m = {}
    for r in rows(G / "DDB-GeneID-UniProt.txt"):
        if len(r) >= 2 and r[0].startswith("DDB") and r[1].startswith("DDB_G"):
            m[r[0].strip()] = r[1].strip()
    return m


def build_extras(ddb2g):
    extras = {}

    def e(g):
        return extras.setdefault(g, {})

    # literature: PMID <tab> DDB_G..._RTE <tab> DDB feature id
    for r in rows(DL / "general" / "DDBID_PMID.txt", skip_header=False):
        if len(r) < 2 or not r[0].strip().isdigit():
            continue
        mm = DDBG.search(r[1])
        if mm:
            e(mm.group(0)).setdefault("pmids", set()).add(r[0].strip())

    # curation status: DDB_G <tab> free-text status
    for r in rows(G / "DDB_G-curation_status.txt", skip_header=False):
        if len(r) >= 2 and r[0].startswith("DDB_G"):
            e(r[0].strip())["curation"] = r[1].strip()

    # alternative transcripts: DDB_G <tab> name <tab> "DDB0..., DDB0..."
    for r in rows(G / "alternative_transcripts.txt"):
        if len(r) >= 3 and r[0].startswith("DDB_G"):
            ts = [t.strip() for t in r[2].split(",") if t.strip()]
            if len(ts) > 1:
                e(r[0].strip())["transcripts"] = ts

    # dictyBase orthologs: source <tab> DDB_G <tab> name <tab> "Sp:ID | Sp:ID"
    for r in rows(G / "ortholog_information.txt"):
        if len(r) >= 4 and r[1].startswith("DDB_G"):
            orths = [o.strip() for o in r[3].split("|") if o.strip()]
            if orths:
                e(r[1].strip()).setdefault("orthologs", {})[r[0].strip()] = orths

    # N-terminal myristoylation: DDB feature id <tab> name <tab> reliability <tab> score
    for r in rows(G / "NMT.txt"):
        if len(r) >= 4:
            g = ddb2g.get(r[0].strip())
            if g:
                e(g)["myristoylation"] = {"reliability": r[2].strip(), "score": r[3].strip()}

    # molecular weight (Da): DDB feature id <tab> DDB_G <tab> MW
    for r in rows(G / "dicty-mw.txt", skip_header=False):
        if len(r) >= 3 and r[1].startswith("DDB_G"):
            try:
                e(r[1].strip())["mw"] = round(float(r[2]), 1)
            except ValueError:
                pass

    # phosphoproteome (AX2, Charest & Firtel): peptides with a phosphosite
    try:
        import openpyxl
        wb = openpyxl.load_workbook(G / "AX2_phosphoproteome.xlsx", read_only=True)
        ws = wb[wb.sheetnames[0]]
        hdr = [str(c or "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        pep_i = next((i for i, h in enumerate(hdr) if h.strip().lower() == "peptide"), 3)
        gene_i = next((i for i, h in enumerate(hdr) if h.strip().lower() == "gene"), 7)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(pep_i, gene_i):
                continue
            gcell = str(row[gene_i] or "")
            pep = str(row[pep_i] or "").strip()
            mm = DDBG.search(gcell)
            if mm and pep:
                p = e(mm.group(0)).setdefault("phospho", {"peptides": set()})
                p["peptides"].add(pep)
    except Exception as exc:  # noqa: BLE001 — phospho is optional
        print(f"  (phospho skipped: {exc})")

    # GO-slim: broad GO categories per gene (mapped to GO-slim.obo), keyed by DDB
    p = DL / "goslim" / "slim_gene_association.ddb"
    if p.exists():
        for r in rows(p, skip_header=False):
            if len(r) > 8 and r[4].startswith("GO:"):
                g = ddb2g.get(r[1].strip())
                if g:
                    e(g).setdefault("goslim", set()).add((r[4].strip(), r[8].strip()))

    # finalize sets -> sorted lists / counts
    for g, d in extras.items():
        if "pmids" in d:
            d["pmids"] = sorted(d["pmids"], key=int)
        if "phospho" in d:
            peps = sorted(d["phospho"]["peptides"])
            d["phospho"] = {"count": len(peps), "peptides": peps[:40]}
        if "goslim" in d:
            d["goslim"] = sorted([g_, a] for g_, a in d["goslim"])
    (ASSETS / "gene_extras.json").write_text(json.dumps(extras, separators=(",", ":")))
    return extras


def build_domains(ddb2g):
    # dictyBase ID, CRC64, Length, Database, Domain ID, Domain Name, Start, End,
    # Score, Status, Date, InterPro ID, InterPro Name
    out = {}
    seen = {}
    for r in rows(G / "Dd_protein_domains.txt"):
        if len(r) < 6:
            continue
        g = ddb2g.get(r[0].strip())
        if not g:
            continue
        db, did, name = r[3].strip(), r[4].strip(), r[5].strip()
        ipr = (r[11].strip() if len(r) > 11 else "")
        ipr_name = (r[12].strip() if len(r) > 12 else "")
        try:
            start, end = int(r[6]), int(r[7])
        except (ValueError, IndexError):
            start = end = None
        key = (g, ipr or did, start)
        if key in seen:
            continue
        seen[key] = True
        out.setdefault(g, []).append({
            "db": db, "id": did, "name": ipr_name or name,
            "start": start, "end": end,
            "interpro": ipr, "interpro_name": ipr_name,
        })
    for g in out:
        out[g].sort(key=lambda d: (d["start"] if d["start"] is not None else 1 << 30))
    (ASSETS / "dictybase_domains.json").write_text(json.dumps(out, separators=(",", ":")))
    return out


def build_promoters(ddb2g):
    """5' flanking (promoter) sequence per gene, from the promoter FASTA. Keyed by
    the DDB feature id in each header (>DDB0…|contig|…); mapped to DDB_G."""
    import zipfile
    zp = DL / "sequence_sets" / "promoter_sequences.zip"
    out = {}
    if not zp.exists():
        (ASSETS / "promoters.json").write_text("{}")
        return out
    with zipfile.ZipFile(zp) as z:
        with z.open("promoter_sequences.fasta") as fh:
            cur, seq = None, []
            for raw in fh:
                line = raw.decode("utf-8", "replace").rstrip("\n")
                if line.startswith(">"):
                    if cur and seq:
                        out[cur] = "".join(seq)
                    ddb0 = line[1:].split("|")[0].strip()
                    cur, seq = ddb2g.get(ddb0), []
                elif cur:
                    seq.append(line.strip())
            if cur and seq:
                out[cur] = "".join(seq)
    (ASSETS / "promoters.json").write_text(json.dumps(out, separators=(",", ":")))
    return out


def build_goslim_terms(extras):
    """Resolve the ~120 distinct GO-slim term ids to names once (via QuickGO),
    mapping secondary/obsolete ids too so the 2006-era slim ids still name-match.
    Written to goslim_terms.json so the client needs no live GO lookups."""
    ids = sorted({go for d in extras.values() for go, _ in d.get("goslim", [])})
    names = {}
    try:
        for i in range(0, len(ids), 100):
            batch = ids[i:i + 100]
            url = ("https://www.ebi.ac.uk/QuickGO/services/ontology/go/terms/"
                   + ",".join(batch))
            req = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": _UA})
            data = json.loads(urllib.request.urlopen(req, timeout=60, context=_SSL).read())
            for t in data.get("results", []):
                if not t.get("name") or t.get("isObsolete"):
                    continue
                names[t["id"]] = t["name"]
                for sid in (t.get("secondaryIds") or []):
                    names.setdefault(sid, t["name"])
    except Exception as exc:  # noqa: BLE001 — best-effort; client falls back to ids
        print(f"  (goslim term names skipped: {exc})")
    (ASSETS / "goslim_terms.json").write_text(json.dumps(names, separators=(",", ":")))
    return names


def build_ontology():
    """Parse the phenotype + anatomy OBO into {term_lower: {...}}."""
    onto = {}
    for fn in ("dicty_phenotypes.obo", "dicty_anatomy.obo"):
        p = DL / "pheno_ontology" / fn
        if not p.exists():
            continue
        term = None
        for line in p.read_text(encoding="utf-8", errors="replace").splitlines():
            line = line.strip()
            if line == "[Term]":
                term = {"id": "", "name": "", "definition": "", "synonyms": [], "parents": []}
            elif term is not None and line.startswith("id:"):
                term["id"] = line[3:].strip()
            elif term is not None and line.startswith("name:"):
                term["name"] = line[5:].strip()
            elif term is not None and line.startswith("def:"):
                m = re.search(r'"([^"]*)"', line)
                term["definition"] = m.group(1) if m else ""
            elif term is not None and line.startswith("synonym:"):
                m = re.search(r'"([^"]*)"', line)
                if m:
                    term["synonyms"].append(m.group(1))
            elif term is not None and line.startswith("is_a:"):
                m = re.search(r"!\s*(.+)$", line)
                if m:
                    term["parents"].append(m.group(1).strip())
            elif line == "" and term is not None and term["name"]:
                onto[term["name"].lower()] = {k: v for k, v in term.items() if k != "name"}
                term = None
    (ASSETS / "phenotype_ontology.json").write_text(json.dumps(onto, separators=(",", ":")))
    return onto


def build_codon_usage():
    """Dictyostelium codon bias -> {amino acid: {codon: fraction}}."""
    text = (G / "Dd_Codon_Bias.txt").read_text(encoding="utf-8", errors="replace")
    # rows look like: CODON <tab> AA <tab> Number <tab> % <tab> %absolute
    counts = {}
    for m in re.finditer(r"\b([ACGTU]{3})\t([A-Za-z*]{1,4})\t(\d+)\t", text):
        codon, aa, n = m.group(1).replace("U", "T"), m.group(2), int(m.group(3))
        counts.setdefault(aa, {})[codon] = counts.get(aa, {}).get(codon, 0) + n
    usage = {}
    for aa, cod in counts.items():
        tot = sum(cod.values()) or 1
        usage[aa] = {c: round(n / tot, 4) for c, n in sorted(cod.items(), key=lambda x: -x[1])}
    (ASSETS / "codon_usage.json").write_text(json.dumps(usage, separators=(",", ":")))
    return usage


def main():
    ddb2g = ddb_to_ddbg()
    print(f"  DDB->DDB_G map: {len(ddb2g)}")
    extras = build_extras(ddb2g)
    lit = sum(1 for d in extras.values() if d.get("pmids"))
    print(f"  gene_extras.json: {len(extras)} genes "
          f"(literature {lit}, "
          f"curation {sum(1 for d in extras.values() if d.get('curation'))}, "
          f"orthologs {sum(1 for d in extras.values() if d.get('orthologs'))}, "
          f"myristoylation {sum(1 for d in extras.values() if d.get('myristoylation'))}, "
          f"phospho {sum(1 for d in extras.values() if d.get('phospho'))}, "
          f"mw {sum(1 for d in extras.values() if d.get('mw'))})")
    dom = build_domains(ddb2g)
    print(f"  dictybase_domains.json: {len(dom)} genes, "
          f"{sum(len(v) for v in dom.values())} domain rows")
    prom = build_promoters(ddb2g)
    print(f"  promoters.json: {len(prom)} genes with a 5' flanking sequence")
    gt = build_goslim_terms(extras)
    print(f"  goslim_terms.json: {len(gt)} GO-slim term names resolved")
    onto = build_ontology()
    print(f"  phenotype_ontology.json: {len(onto)} terms")
    cod = build_codon_usage()
    print(f"  codon_usage.json: {len(cod)} amino acids")


if __name__ == "__main__":
    main()
